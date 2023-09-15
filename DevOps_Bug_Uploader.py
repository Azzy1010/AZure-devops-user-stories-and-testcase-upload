import openpyxl
import time
import csv
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import tkinter as tk
from tkinter import filedialog
import pyautogui
from selenium.webdriver.chrome.options import Options

def browser_open():
    pyautogui.hotkey('win')
    time.sleep(.5)
    pyautogui.write('cmd')
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.write('cd "C:\Program Files\Google\Chrome\Application"')
    pyautogui.press('enter')
    pyautogui.write('chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\Chromedata"')
    pyautogui.press('enter')
    pyautogui.write('exit')
    pyautogui.press('enter')

def image_select():
    global fi
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def copy_image(image_file_loc, filename_to_upload):
    pyautogui.hotkey('ctrl', 'l')
    pyautogui.write(image_file_loc)
    pyautogui.hotkey('enter')
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'f')
    pyautogui.write(f'{filename_to_upload}')
    pyautogui.press('down')
    time.sleep(1)
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('enter')
    time.sleep(3)
def file_excel():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

pyautogui.hotkey('ctrl', 'D')
time.sleep(1)

pyautogui.alert(text='Please Select Images folder location', title='Image Selection')

a = image_select().split('/')
value = len(a) - 1
a.pop(value)
image_file_loc = '/'.join(a)

pyautogui.alert(text='Please Select Excel file', title='Excel Selection')
excel_loc = file_excel()

browser_open()
options = Options()
options.add_experimental_option("debuggerAddress", "localhost:9222")
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.get('https://dev.azure.com/ObeikanDigitalSol/O3Scrum/_sprints/taskboard/O3Scrum%20Team/O3Scrum/Sprint%203')

driver.implicitly_wait(4)
pyautogui.hotkey('esc')
time.sleep(1)
dataframe = openpyxl.load_workbook(f"{excel_loc}")
dataframe1 = dataframe.active
global Summary,Description,Expected,Actual,priority,severity,repro,image_name,sp_no,assignee,state
count = 0


for row in range(1, dataframe1.max_row):
    exc_turn = 0
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if exc_turn == 0:
            Summary = col[row].value
        elif exc_turn == 1:
            Description = col[row].value
        elif exc_turn == 2:
            Expected = col[row].value
        elif exc_turn == 3:
            Actual = col[row].value
        elif exc_turn == 4:
            priority = col[row].value
        elif exc_turn == 5:
            severity = col[row].value
        elif exc_turn == 6:
            repro = col[row].value
        elif exc_turn == 7:
            image_name = col[row].value
        elif exc_turn == 8:
            sp_no = col[row].value
        elif exc_turn == 9:
            assignee = col[row].value
        elif exc_turn == 10:
            state = col[row].value
        exc_turn = exc_turn + 1
    # print(Summary,Description,Expected,Actual,priority,severity,repro,image_name,sp_no,assignee,state)
    if Summary is None:
        print(f'{count + 1}. [Skipped] Title missing')
        count = count + 1
        pass
    else:
        driver.find_element(By.XPATH,
                            f'//*[contains(@class,"ms-CommandBarItem-commandText item") and contains(text(),"Sprint")]').click()
        pyautogui.write(f'Sprint {sp_no}')
        time.sleep(0.5)
        for i in range(2): pyautogui.press('down')
        time.sleep(1)
        pyautogui.press('enter')
        driver.implicitly_wait(3)
        try:
            driver.find_element(By.XPATH, '//*[@aria-label="Filter by keyword"]').click()
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            time.sleep(0.5)
            driver.find_element(By.XPATH, "//*[@aria-label='Filter']").click()
        except:
            pass
        driver.find_element(By.XPATH, "//*[contains(@name,'New Work Item')]").click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, "//*[contains(@name,'Bug')]").click()
        time.sleep(0.5)
        pyautogui.write(f'{Summary}')
        time.sleep(1)
        pyautogui.hotkey('enter')
        time.sleep(2)
        try:
            click_bug = driver.find_element(By.XPATH, f'//*[contains(@aria-label,"Bug, {Summary}")]')
            bug_id = click_bug.get_attribute('id').split('-')[1]
            if len(bug_id) == 0:
                assert False
            else:
                pass
        except:
            driver.implicitly_wait(7)
            click_bug = driver.find_element(By.XPATH, f'//*[contains(@aria-label,"Bug, {Summary}")]')
            bug_id = click_bug.get_attribute('id').split('-')[1]
        driver.find_element(By.XPATH, f'//*[@id="tile-{bug_id}"]/div/div[4]').click()
        time.sleep(0.5)
        pyautogui.press('down')
        pyautogui.hotkey('enter')
        driver.implicitly_wait(5)
        try:
            if priority is None:
                pass
            else:
                Priority_box = driver.find_element(By.XPATH,
                                                   '//*[contains(@aria-label,"Priority") and contains(@role,"combobox")]')
                driver.find_element(By.XPATH, f'{Priority_box}').click()
                driver.find_element(By.XPATH, f'{Priority_box}').clear()
                driver.find_element(By.XPATH, f'{Priority_box}').send_keys(priority)
                time.sleep(1)
            if severity is None:
                pass
            else:
                severity_box = '//*[contains(@aria-label,"Severity") and contains(@role,"combobox")]'
                driver.find_element(By.XPATH, f'{severity_box}').click()
                driver.find_element(By.XPATH, f'{severity_box}').clear()
                driver.find_element(By.XPATH, f'{severity_box}').send_keys(f'{severity}' + Keys.ARROW_RIGHT)
                time.sleep(1)
                pyautogui.hotkey('enter')
            if state is None:
                pass
            else:
                state_box = '//*[@aria-label="State Field"]'
                driver.find_element(By.XPATH, state_box).click()
                driver.implicitly_wait(1)
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.5)
                pyautogui.write(f'{state}')
                pyautogui.press('enter')
                time.sleep(0.5)
            if Description is None:
                pass
            else:
                defect_desc = "//*[contains(@aria-label,'Defect Description') and contains(@data-placeholder,'Click to add Defect Description')]"
                driver.find_element(By.XPATH, f"{defect_desc}").click()
                driver.find_element(By.XPATH, f"{defect_desc}").clear()
                driver.find_element(By.XPATH, f"{defect_desc}").send_keys(f'{Description}')
                time.sleep(2)
            if repro is None:
                pass
            else:
                repro_steps = "//*[contains(@data-placeholder,'Click to add Repro Steps')]"
                driver.find_element(By.XPATH, f"{repro_steps}").click()
                driver.find_element(By.XPATH, f"{repro_steps}").send_keys(f'{repro}')
                time.sleep(2)
            if Expected is None:
                pass
            else:
                Expected_result = (
                    "//*[contains(@role,'textbox') and contains(@data-placeholder,'Click to add Expected Result')]")
                driver.find_element(By.XPATH, f"{Expected_result}").click()
                driver.find_element(By.XPATH, f"{Expected_result}").clear()
                driver.find_element(By.XPATH, f"{Expected_result}").send_keys(f'{Expected}')
                time.sleep(2)
            if Actual is None:
                pass
            else:
                Actual_result = (
                    "//*[contains(@role,'textbox') and contains(@data-placeholder,'Click to add Actual Result')]")
                driver.find_element(By.XPATH, f"{Actual_result}").click()
                driver.find_element(By.XPATH, f"{Actual_result}").clear()
                driver.find_element(By.XPATH, f"{Actual_result}").send_keys(f'{Actual}')
                time.sleep(2)
            if image_name is None:
                pass
            else:
                driver.find_element(By.XPATH,
                                    '//*[contains(@aria-label,"Attachments") and @aria-posinset="4"]').click()
                time.sleep(5)
                driver.find_element(By.XPATH,
                                        '//*[contains(@class,"ms-Button work-item-zero-cta ms-Button--")]').click()
                copy_image(image_file_loc, int(image_name))
            driver.find_element(By.XPATH, "//*[contains(@role,'button') and contains(text(),'Save')]").click()
            driver.implicitly_wait(5)
            if assignee is None:
                pass
            else:
                try:
                    click_bug = driver.find_element(By.XPATH, f'//*[contains(@aria-label,"Bug, {Summary}")]')
                    bug_id = click_bug.get_attribute('id').split('-')[1]
                    if len(bug_id) == 0:
                        assert False
                    else:
                        pass
                except:
                    time.sleep(7)
                    click_bug = driver.find_element(By.XPATH, f'//*[contains(@aria-label,"Bug, {Summary}")]')
                    bug_id = click_bug.get_attribute('id').split('-')[1]
                driver.find_element(By.XPATH, f'//*[@id="tile-{bug_id}"]/div/div[2]/div[2]/div').click()
                time.sleep(0.5)
                pyautogui.write(f'{assignee}')
                pyautogui.press('enter')
                time.sleep(0.5)
                pyautogui.press('enter')
            print(f'{count + 1}. [Success] {Summary}')
        except:
            print(f'{count + 1}. [Failed] {Summary}')
        count = count + 1
print('Process completed!')
driver.close()

