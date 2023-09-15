import time
import csv
from selenium.webdriver.common.by import By
from selenium import webdriver
import tkinter as tk
from tkinter import filedialog
import pyautogui
from selenium.webdriver.chrome.options import Options
import openpyxl


def browser_open():
    pyautogui.hotkey('win')
    time.sleep(.5)
    pyautogui.write('cmd')
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.write('cd "C:\Program Files\Google\Chrome\Application"')
    pyautogui.press('enter')
    pyautogui.write('chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\Chromedata"')
    pyautogui.press('enter')
    pyautogui.write('exit')
    pyautogui.press('enter')


def file_csv():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path


def test_cases():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path


def rows_counter(file_loc1):
    dataframe = openpyxl.load_workbook(f"{file_loc1}")
    dataframe1 = dataframe.active
    global name, s_name, t_name, f_name, fi_name
    rows_count = 0
    for row in range(1, dataframe1.max_row):
        rows_count = rows_count + 1
    print(f'User Story(s) found -- {rows_count}')


def chrome_config():
    global driver
    browser_open()
    options = Options()
    options.add_experimental_option("debuggerAddress", "localhost:9222")
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()




def user_stories_creation():
    error = 0
    turn = 0
    dataframe = openpyxl.load_workbook(f"{file_loc}")
    dataframe1 = dataframe.active
    global title, priority_val, who_val, wants_to_do_val, so_that_val, given_val, when_val, then_val, feature
    for row in range(1, dataframe1.max_row):
        driver.get('https://dev.azure.com/ObeikanDigitalSol/O3Scrum/_backlogs/backlog/O3Scrum%20Team/Features')
        time.sleep(3)
        pyautogui.hotkey('esc')
        time.sleep(0.5)
        exc_turn = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if exc_turn == 0:
                title = str(col[row].value)
            elif exc_turn == 1:
                feature = col[row].value
            elif exc_turn == 2:
                who_val = str(col[row].value)
            elif exc_turn == 3:
                wants_to_do_val = str(col[row].value)
            elif exc_turn == 4:
                so_that_val = str(col[row].value)
            elif exc_turn == 5:
                given_val = str(col[row].value)
            elif exc_turn == 6:
                when_val = str(col[row].value)
            elif exc_turn == 7:
                then_val = str(col[row].value)
            exc_turn = exc_turn + 1
        try:
            driver.find_element(By.XPATH,
                                f"//*[contains(@class,'work-item-title-link') and contains(text(),'{feature}')]").click()
        except:
            print("failed to find EPIC")
            driver.quit()
            exit()
        driver.implicitly_wait(7)
        driver.find_element(By.XPATH, "//*[@class='menu-item icon-only' and @aria-posinset='7']").click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//*[text()='New linked work item']").click()
        time.sleep(1)
        combo_box = "//*[contains(@class , 'ms-ComboBox-Input')]"
        driver.find_element(By.XPATH, combo_box).click()
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('backspace')
        driver.find_element(By.XPATH, combo_box).send_keys('Child')
        time.sleep(1)
        work_item = "//*[contains(@id,'work-item-types_txt')]"
        driver.find_element(By.XPATH, work_item).click()
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('backspace')
        driver.find_element(By.XPATH, work_item).send_keys('Product Backlog Item')
        pyautogui.hotkey('enter')
        tc_title = "//*[contains(@id,'dialog-label')]"
        driver.find_element(By.XPATH, tc_title).click()
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('backspace')
        driver.find_element(By.XPATH, tc_title).send_keys(title)
        time.sleep(1)
        time.sleep(0.5)
        pyautogui.press('down')
        pyautogui.hotkey('enter')
        time.sleep(5)
        try:
            who = '//*[@aria-label="Who"]'
            driver.find_element(By.XPATH, who).click()
            driver.find_element(By.XPATH, who).clear()
            driver.find_element(By.XPATH, who).send_keys(f'{who_val}')
            time.sleep(0.5)
            Wants_to_do = '//*[@aria-label="Wants to do"]'
            driver.find_element(By.XPATH, Wants_to_do).click()
            driver.find_element(By.XPATH, Wants_to_do).clear()
            driver.find_element(By.XPATH, Wants_to_do).send_keys(f'{wants_to_do_val}')
            time.sleep(0.5)
            so_that = '//*[@aria-label="so that"]'
            driver.find_element(By.XPATH, so_that).click()
            driver.find_element(By.XPATH, so_that).clear()
            driver.find_element(By.XPATH, so_that).send_keys(f'{so_that_val}')
            time.sleep(0.5)
            Given = '//*[@aria-label="Given"]'
            driver.find_element(By.XPATH, Given).click()
            driver.find_element(By.XPATH, Given).clear()
            driver.find_element(By.XPATH, Given).send_keys(f'{given_val}')
            time.sleep(0.5)
            When = '//*[@aria-label="When"]'
            driver.find_element(By.XPATH, When).click()
            driver.find_element(By.XPATH, When).clear()
            driver.find_element(By.XPATH, When).send_keys(f'{when_val}')
            time.sleep(0.5)
            Then = '//*[@aria-label="Then"]'
            driver.find_element(By.XPATH, Then).click()
            driver.find_element(By.XPATH, Then).clear()
            driver.find_element(By.XPATH, Then).send_keys(f'{then_val}')
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 's')
            time.sleep(2)
            print(f'{turn + 1}. {title} user story created Successfully')
        except:
            print(f'{turn + 1}. Failed to create {title} user story')
        turn = turn + 1
    print('User stories Created Successfully, proceeding with the testcases...')
    driver.quit()

def testcase_creation():
    driver.get('https://dev.azure.com/ObeikanDigitalSol/O3Scrum/_sprints/taskboard/O3Scrum%20Team/O3Scrum/Sprint%203')
    time.sleep(3)
    pyautogui.press('esc')
    dataframe = openpyxl.load_workbook(f"{test_Cases_loc}")
    dataframe1 = dataframe.active
    global title_US, test_c_name, pre_req_data, testcase_data, steps_data, expected_data, scenarios_data
    tc_turn = 0
    for row in range(1, dataframe1.max_row):
        exc_turn = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if exc_turn == 0:
                title_US = str(col[row].value)
            elif exc_turn == 1:
                test_c_name = col[row].value
            elif exc_turn == 2:
                pre_req_data = str(col[row].value)
            elif exc_turn == 3:
                testcase_data = str(col[row].value)
            elif exc_turn == 4:
                steps_data = str(col[row].value)
            elif exc_turn == 5:
                expected_data = str(col[row].value)
            elif exc_turn == 6:
                scenarios_data = str(col[row].value)
            exc_turn = exc_turn + 1
        try:
            try:
                driver.find_element(By.XPATH, '//*[@aria-label="Filter by keyword"]').click()
                time.sleep(1)
                pyautogui.hotkey('ctrl', 'a')
                pyautogui.press('backspace')
                time.sleep(0.5)
            except:
                driver.find_element(By.XPATH, '//*[@data-command-key="work-item-filter-bar"]').click()
                time.sleep(0.5)
                driver.find_element(By.XPATH, '//*[@aria-label="Filter by keyword"]').click()
                time.sleep(0.5)
                pyautogui.hotkey('ctrl', 'a')
                pyautogui.press('backspace')
                time.sleep(0.5)
            driver.find_element(By.XPATH, '//*[@aria-label="Filter by keyword"]').send_keys(f'{title_US}')
            time.sleep(1)
            search_check = driver.find_elements(By.XPATH, '//*[@class="clickable-title"]')
            if len(search_check) > 2:
                pyautogui.alert(text=f'Please select {title_US} and press OK')
            else:
                pyautogui.hotkey('enter')
                try:
                    driver.find_element(By.XPATH,
                                        f'//*[contains(text(),"{title_US}") and contains(@class,"clickable-title")]').click()
                except:
                    e = 'User story not created'
            time.sleep(3)
            driver.find_element(By.XPATH, "//*[contains(@aria-label,'Links') and (@aria-posinset='3')]").click()
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@class='menu-item icon-only' and @aria-posinset='7']").click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, "//*[text()='New linked work item']").click()
            time.sleep(1)
            combo_box = "//*[contains(@class , 'ms-ComboBox-Input')]"
            driver.find_element(By.XPATH, combo_box).click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            driver.find_element(By.XPATH, combo_box).send_keys('Child')
            time.sleep(1)
            work_item = "//*[contains(@id,'work-item-types_txt')]"
            driver.find_element(By.XPATH, work_item).click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            driver.find_element(By.XPATH, work_item).send_keys('Test Case')
            pyautogui.hotkey('enter')
            tc_title = "//*[contains(@id,'dialog-label')]"
            driver.find_element(By.XPATH, tc_title).click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            driver.find_element(By.XPATH, tc_title).send_keys(test_c_name)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id='ok']").click()
            time.sleep(2)
            pre_rep = "//*[contains(@data-placeholder,'Click to add PreRequisites')]"
            driver.find_element(By.XPATH, pre_rep).click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            driver.find_element(By.XPATH, pre_rep).send_keys(pre_req_data)
            time.sleep(1)
            test_data = "//*[contains(@data-placeholder,'Click to add Test Data')]"
            if testcase_data == 'None':
                driver.find_element(By.XPATH, test_data).click()
                pyautogui.hotkey('ctrl', 'a')
                pass
            else:
                driver.find_element(By.XPATH, test_data).click()
                pyautogui.hotkey('ctrl', 'a')
                pyautogui.press(test_data)
                driver.find_element(By.XPATH, test_data).send_keys(testcase_data)
                time.sleep(0.5)
            driver.find_element(By.XPATH, "//p[contains(text(),'Click or type here to add a step')]").click()
            time.sleep(1)
            element = driver.find_element(By.XPATH,
                                          "//*[contains(@class,'grid-row grid-row-normal') and contains(@role,'button')]")
            value = element.get_attribute('id')
            new_value = int(value.split('_')[2])
            time.sleep(1)
            for i in range(len(steps_data.split('\n'))):
                steps = f"//*[@id='row_vss_{new_value}_{i}']/div[2]"
                if i > 0:
                    driver.find_element(By.XPATH, steps).click()
                pyautogui.write(steps_data.split('\n')[i])
                time.sleep(0.5)
                expected = f"//*[@id='row_vss_{new_value}_{i}']/div[3]"
                driver.find_element(By.XPATH, expected).click()
                pyautogui.write(expected_data.split('\n')[i])
                if i < int(len(steps_data.split('\n'))):
                    pass
                else:
                    pyautogui.hotkey('enter')
            time.sleep(1)
            driver.find_element(By.XPATH, '//*[@aria-label="Summary"]').click()
            time.sleep(0.5)
            scenario = "//*[@aria-label='Test Scenario']"
            driver.find_element(By.XPATH, scenario).click()
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            driver.find_element(By.XPATH, scenario).send_keys(scenarios_data)
            time.sleep(0.5)
            # driver.find_element(By.XPATH,
            #                     "//*[contains(@command,'save-and-close-work-item') and contains(@aria-setsize,'5')]").click()
            time.sleep(8)
            pyautogui.hotkey('esc')
            time.sleep(0.5)
            print(f'{tc_turn + 1}. {title_US} testcase created Successfully')
            tc_turn = tc_turn + 1
        except:
            print(f'{tc_turn + 1}. Failed --- {title_US}, {test_c_name}, Please check')
            tc_turn = tc_turn + 1
    print('All test cases added successfully')
    driver.close()

choice = int(pyautogui.prompt(
    text='Please press 1 for [User Stories] 2 for [testcases] and 3 for both (user stories --> testcases'))





if choice == 1:
    pyautogui.alert(text= 'Select User Stories file')
    file_loc = file_csv()
    rows_counter(file_loc)
    time.sleep(1)
    chrome_config()
    user_stories_creation()
elif choice == 2:
    pyautogui.alert(text='Select testcases file')
    test_Cases_loc = test_cases()
    time.sleep(1)
    chrome_config()
    testcase_creation()
elif choice == 3:
    pyautogui.alert(text='Select User Stories file')
    file_loc = file_csv()
    pyautogui.alert(text='Select testcases file')
    test_Cases_loc = test_cases()
    rows_counter(file_loc)
    time.sleep(1)
    chrome_config()
    user_stories_creation()
    testcase_creation()
else:
    pyautogui.alert(text = 'Wrong choice entered!\nValue should be <= 4',title = 'Alert!')

